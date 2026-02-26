import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference

def generate_sample_data(filename="input_esg_data.xlsx"):
    data = {
        'Company': ['Firma A', 'Firma B', 'Firma C', 'Firma D'],
        'Revenue': [200.0, 450.0, 100.0, 600.0],
        'Emissions_Scope1': [5000, 30000, 2500, 18000],
        'Emissions_Scope2': [2000, 10000, 800, 6000],
        'Emissions_Scope3': [15000, 60000, 5000, 25000],  # opcjonalne Scope 3
        'Employees': [150, 1000, 60, 1500],
        'Sector': ['Technologia', 'Produkcja', 'Usługi', 'Transport']
    }
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

def create_esrs_report(input_file, output_file):
    df = pd.read_excel(input_file)
    
    # Obliczenia podstawowe
    df['Total_Emissions_S1_S2'] = df['Emissions_Scope1'] + df['Emissions_Scope2']
    df['Total_Emissions_All'] = df['Total_Emissions_S1_S2'] + df.get('Emissions_Scope3', 0)
    df['Carbon_Intensity_S1_S2'] = round(df['Total_Emissions_S1_S2'] / df['Revenue'], 2)
    df['Carbon_Intensity_All'] = round(df['Total_Emissions_All'] / df['Revenue'], 2)
    df['Employee_Intensity'] = round(df['Employees'] / df['Revenue'], 2)
    
    # Dynamiczne progi per sektor (przykładowe – możesz rozbudować)
    def get_carbon_color(val, sector):
        if sector in ['Technologia', 'Usługi']:
            if val < 40: return 'green'
            elif val <= 100: return 'yellow'
            else: return 'red'
        else:  # Produkcja, Transport itp.
            if val < 150: return 'green'
            elif val <= 400: return 'yellow'
            else: return 'red'
    
    columns = ['Company', 'Sector', 'Revenue', 'Employees', 'Emissions_Scope1',
               'Emissions_Scope2', 'Emissions_Scope3', 'Total_Emissions_All',
               'Carbon_Intensity_S1_S2', 'Carbon_Intensity_All', 'Employee_Intensity']
    df = df[columns]
    
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Dane ESG (ESRS)"
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)
    
    # Formatowanie nagłówków
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
    
    # Zamrożenie nagłówka
    ws_data.freeze_panes = "A2"
    
    # Szerokość kolumn
    for column_cells in ws_data.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws_data.column_dimensions[column_cells[0].column_letter].width = length + 2
    
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Kolorowanie Carbon_Intensity_All (kolumna 10)
    for row in range(2, ws_data.max_row + 1):
        sector = ws_data.cell(row=row, column=2).value
        val = ws_data.cell(row=row, column=10).value
        if isinstance(val, (int, float)):
            color = get_carbon_color(val, sector)
            fill = green if color == 'green' else yellow if color == 'yellow' else red
            ws_data.cell(row=row, column=10).fill = fill
    
    # Kolorowanie Employee_Intensity (kolumna 11)
    for row in range(2, ws_data.max_row + 1):
        val = ws_data.cell(row=row, column=11).value
        if isinstance(val, (int, float)):
            if val < 1:
                ws_data.cell(row=row, column=11).fill = green
            elif 1 <= val <= 3:
                ws_data.cell(row=row, column=11).fill = yellow
            else:
                ws_data.cell(row=row, column=11).fill = red
    
    # Dashboard z dwoma wykresami
    ws_charts = wb.create_sheet(title="Wykresy ESRS")
    
    chart1 = BarChart()
    chart1.title = "Carbon Intensity – Benchmark Firm 2026 (tCO₂e / mln EUR)"
    chart1.y_axis.title = "Intensywność węglowa"
    chart1.x_axis.title = "Firma"
    data1 = Reference(ws_data, min_col=10, min_row=1, max_row=ws_data.max_row)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    ws_charts.add_chart(chart1, "A1")
    
    chart2 = BarChart()
    chart2.title = "Employee Intensity (Pracownicy / mln EUR)"
    chart2.y_axis.title = "Intensywność zatrudnienia"
    chart2.x_axis.title = "Firma"
    data2 = Reference(ws_data, min_col=11, min_row=1, max_row=ws_data.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws_charts.add_chart(chart2, "K1")
    
    wb.save(output_file)

if __name__ == "__main__":
    input_file = "input_esg_data.xlsx"
    output_file = "esrs_esg_report_v2.1.xlsx"
    generate_sample_data(input_file)
    create_esrs_report(input_file, output_file)
